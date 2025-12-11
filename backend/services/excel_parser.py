from dotenv import load_dotenv
load_dotenv()

import os
import json
import time
import google.generativeai as genai
from typing import Dict, List, Any
from datetime import datetime

# Configure Gemini
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

EXCEL_EXTRACTION_PROMPT = """
You are an expert at extracting invoice data from Excel spreadsheets.

Analyze this Excel data and extract ALL invoice line items AND summary information. Return ONLY valid JSON with this structure:

{
  "invoices": [
    {
      "serial_number": "invoice/serial number",
      "customer_name": "customer/party name or company name",
      "product_name": "product/item name",
      "quantity": number,
      "tax": tax amount as number (if tax is percentage, calculate actual amount),
      "total_amount": total amount including tax as number,
      "date": "YYYY-MM-DD or MISSING",
      "discount": discount amount or 0,
      "payment_mode": "payment method or MISSING",
      "notes": "status/notes or MISSING"
    }
  ],
  "products": [
    {
      "name": "product name",
      "quantity": total quantity across all invoices,
      "unit_price": price per unit,
      "tax": total tax amount,
      "price_with_tax": total with tax,
      "discount": total discount or 0,
      "sku": "SKU or MISSING"
    }
  ],
  "customers": [
    {
      "customer_name": "customer/party name",
      "phone_number": "phone or MISSING",
      "total_purchase_amount": total amount,
      "email": "email or MISSING",
      "address": "address or MISSING"
    }
  ],
  "summary": {
    "total_quantity": total quantity from all invoices,
    "total_amount": grand total amount,
    "cgst": CGST amount if present,
    "sgst": SGST amount if present,
    "igst": IGST amount if present,
    "net_amount": net amount before tax,
    "total_tax": total tax amount,
    "extra_discount": extra discount if any,
    "round_off": round off amount if any
  }
}

CRITICAL RULES:
1. Extract EVERY row that represents an invoice line item
2. Skip summary rows (like "Totals", "Grand Total") from invoices BUT extract their values for the summary section
3. Skip completely empty rows
4. Each product in an invoice should be a separate invoice entry
5. Use "MISSING" for fields that are not available, use 0 for numeric fields that are not available
6. Convert all numbers to numeric types (not strings)
7. Aggregate products and customers correctly
8. Look for summary rows at the bottom with tax breakdowns (CGST, SGST, IGST)
9. Return ONLY JSON, no markdown, no explanations
"""

def parse_excel(file_path: str) -> dict:
    """Parse Excel with AI fallback to manual parsing"""
    try:
        # Try AI parsing with retry logic
        return parse_excel_with_ai(file_path)
    except Exception as e:
        error_msg = str(e).lower()
        
        # If rate limit or quota error, fall back to manual parsing
        if 'quota' in error_msg or 'rate limit' in error_msg or '429' in error_msg:
            print("DEBUG: Rate limit hit, falling back to manual parsing...")
            return parse_excel_manual(file_path)
        
        # For other errors, try manual parsing as fallback
        print(f"DEBUG: AI parsing failed ({str(e)}), trying manual parsing...")
        try:
            return parse_excel_manual(file_path)
        except Exception as manual_error:
            raise Exception(f"Both AI and manual parsing failed. AI: {str(e)}, Manual: {str(manual_error)}")


def parse_excel_with_ai(file_path: str, max_retries: int = 2) -> dict:
    """Use Gemini AI to parse Excel file with retry logic"""
    if not GEMINI_API_KEY:
        raise Exception("GEMINI_API_KEY not set")
    
    # Convert Excel to text
    excel_text = convert_excel_to_text(file_path)
    print(f"DEBUG: Excel text preview:\n{excel_text[:500]}...")
    
    # Try different models in order of preference
    models_to_try = [
        'gemini-1.5-flash',  # Better rate limits than 2.5
        'gemini-1.5-pro',
        'gemini-2.0-flash-exp'
    ]
    
    last_error = None
    
    for model_name in models_to_try:
        for attempt in range(max_retries):
            try:
                print(f"DEBUG: Trying {model_name} (attempt {attempt + 1}/{max_retries})...")
                
                model = genai.GenerativeModel(model_name)
                response = model.generate_content([EXCEL_EXTRACTION_PROMPT, excel_text])
                
                response_text = response.text.strip()
                response_text = clean_json_response(response_text)
                
                extracted_data = json.loads(response_text)
                extracted_data = validate_and_normalize(extracted_data)
                
                print(f"DEBUG: ✓ Successfully extracted with {model_name}")
                print(f"DEBUG: {len(extracted_data.get('invoices', []))} invoices, {len(extracted_data.get('products', []))} products")
                
                return extracted_data
                
            except Exception as e:
                last_error = e
                error_msg = str(e).lower()
                
                # If rate limit, wait and retry
                if 'quota' in error_msg or 'rate limit' in error_msg or '429' in error_msg:
                    if attempt < max_retries - 1:
                        wait_time = (2 ** attempt) * 5  # Exponential backoff: 5s, 10s
                        print(f"DEBUG: Rate limit hit, waiting {wait_time}s before retry...")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"DEBUG: Rate limit persists after retries, trying next model...")
                        break
                
                # For other errors, try next model immediately
                print(f"DEBUG: Error with {model_name}: {str(e)}")
                break
    
    # All models failed
    raise last_error if last_error else Exception("AI parsing failed")


def parse_excel_manual(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """Manual Excel parsing as fallback"""
    import openpyxl
    
    print("DEBUG: Using manual Excel parsing...")
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Get headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        # Header mappings
        header_map = {
            'serial number': 'serial_number',
            'serial no': 'serial_number',
            'invoice number': 'serial_number',
            'invoice no': 'serial_number',
            'customer name': 'customer_name',
            'customer': 'customer_name',
            'party name': 'customer_name',
            'party company name': 'customer_company',
            'product name': 'product_name',
            'product': 'product_name',
            'item': 'product_name',
            'quantity': 'quantity',
            'qty': 'quantity',
            'tax': 'tax',
            'tax (%)': 'tax_percent',
            'total': 'total_amount',
            'total amount': 'total_amount',
            'item total amount': 'total_amount',
            'price with tax': 'total_amount',
            'date': 'date',
            'invoice date': 'date',
            'price': 'unit_price',
            'unit price': 'unit_price',
            'discount': 'discount',
            'item discount': 'discount',
            'payment mode': 'payment_mode',
            'status': 'status',
        }
        
        # Normalize headers
        normalized_headers = [header_map.get(h, h.replace(' ', '_')) for h in headers]
        
        print(f"DEBUG: Headers: {normalized_headers}")
        
        # Extract data
        invoices = []
        products_map = {}
        customers_map = {}
        summary = {
            'total_quantity': 0,
            'total_amount': 0,
            'cgst': 0,
            'sgst': 0,
            'igst': 0,
            'net_amount': 0,
            'total_tax': 0,
            'extra_discount': 0,
            'round_off': 0
        }
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(normalized_headers):
                    row_data[normalized_headers[idx]] = value
            
            serial = row_data.get('serial_number')
            
            # Skip totals/summary rows
            if serial and str(serial).strip().lower() in ['totals', 'total', 'none', 'summary']:
                continue
            
            product = row_data.get('product_name')
            customer = row_data.get('customer_name') or row_data.get('customer_company')
            
            if not serial and not product:
                continue
            
            # Get values
            serial_number = str(serial).strip() if serial else f'INV-{row_idx}'
            customer_name = str(customer).strip() if customer and str(customer).strip() else 'MISSING'
            product_name = str(product).strip() if product and str(product).strip() else 'MISSING'
            
            qty = safe_int(row_data.get('quantity', 1))
            if qty == 0:
                qty = 1
            
            total_amount = safe_float(row_data.get('total_amount', 0))
            tax = safe_float(row_data.get('tax', 0))
            tax_percent = safe_float(row_data.get('tax_percent', 0))
            
            # Calculate tax from percentage if needed
            if tax == 0 and tax_percent > 0 and total_amount > 0:
                amount_before_tax = total_amount / (1 + tax_percent / 100)
                tax = total_amount - amount_before_tax
            
            discount = safe_float(row_data.get('discount', 0))
            unit_price = safe_float(row_data.get('unit_price', 0))
            
            if unit_price == 0 and total_amount > 0 and qty > 0:
                unit_price = (total_amount - tax) / qty
            
            # Create invoice
            invoice = {
                'serial_number': serial_number,
                'customer_name': customer_name,
                'product_name': product_name,
                'quantity': qty,
                'tax': tax,
                'total_amount': total_amount,
                'date': format_date(row_data.get('date')),
                'discount': discount,
                'payment_mode': str(row_data.get('payment_mode', 'MISSING')),
                'notes': str(row_data.get('status', 'MISSING'))
            }
            invoices.append(invoice)
            
            # Update summary
            summary['total_quantity'] += qty
            summary['total_amount'] += total_amount
            summary['total_tax'] += tax
            
            # Aggregate products
            if product_name != 'MISSING':
                if product_name not in products_map:
                    products_map[product_name] = {
                        'name': product_name,
                        'quantity': qty,
                        'unit_price': unit_price,
                        'tax': tax,
                        'price_with_tax': total_amount,
                        'discount': discount,
                        'sku': 'MISSING'
                    }
                else:
                    products_map[product_name]['quantity'] += qty
                    products_map[product_name]['price_with_tax'] += total_amount
                    products_map[product_name]['tax'] += tax
            
            # Aggregate customers
            if customer_name != 'MISSING':
                if customer_name not in customers_map:
                    customers_map[customer_name] = {
                        'customer_name': customer_name,
                        'phone_number': 'MISSING',
                        'total_purchase_amount': total_amount,
                        'email': 'MISSING',
                        'address': 'MISSING'
                    }
                else:
                    customers_map[customer_name]['total_purchase_amount'] += total_amount
        
        summary['net_amount'] = summary['total_amount'] - summary['total_tax']
        
        print(f"DEBUG: Manual parsing complete - {len(invoices)} invoices")
        
        return {
            'invoices': invoices,
            'products': list(products_map.values()),
            'customers': list(customers_map.values()),
            'summary': summary
        }
        
    except Exception as e:
        raise Exception(f"Manual Excel parsing error: {str(e)}")


def convert_excel_to_text(file_path: str) -> str:
    """Convert Excel to structured text for AI"""
    import openpyxl
    
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    text_lines = ["EXCEL DATA:", "=" * 80]
    
    headers = [str(cell.value) for cell in sheet[1] if cell.value]
    text_lines.append("HEADERS: " + " | ".join(headers))
    text_lines.append("-" * 80)
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        row_data = [f"{headers[idx]}: {value}" for idx, value in enumerate(row) if idx < len(headers) and value]
        if row_data:
            text_lines.append(f"ROW {row_idx}: " + " | ".join(row_data))
    
    return "\n".join(text_lines)


def clean_json_response(text: str) -> str:
    """Clean AI response"""
    if '```json' in text:
        text = text.split('```json')[1].split('```')[0]
    elif '```' in text:
        text = text.split('```')[1].split('```')[0]
    return text.strip()


def validate_and_normalize(data: dict) -> dict:
    """Validate and normalize extracted data"""
    data.setdefault('invoices', [])
    data.setdefault('products', [])
    data.setdefault('customers', [])
    data.setdefault('summary', {})
    
    for invoice in data['invoices']:
        invoice['serial_number'] = str(invoice.get('serial_number', 'MISSING'))
        invoice['customer_name'] = str(invoice.get('customer_name', 'MISSING'))
        invoice['product_name'] = str(invoice.get('product_name', 'MISSING'))
        invoice['quantity'] = safe_int(invoice.get('quantity', 1))
        invoice['tax'] = safe_float(invoice.get('tax', 0))
        invoice['total_amount'] = safe_float(invoice.get('total_amount', 0))
        invoice['date'] = str(invoice.get('date', 'MISSING'))
        invoice['discount'] = safe_float(invoice.get('discount', 0))
        invoice['payment_mode'] = str(invoice.get('payment_mode', 'MISSING'))
        invoice['notes'] = str(invoice.get('notes', 'MISSING'))
    
    for product in data['products']:
        product['name'] = str(product.get('name', 'MISSING'))
        product['quantity'] = safe_int(product.get('quantity', 0))
        product['unit_price'] = safe_float(product.get('unit_price', 0))
        product['tax'] = safe_float(product.get('tax', 0))
        product['price_with_tax'] = safe_float(product.get('price_with_tax', 0))
        product['discount'] = safe_float(product.get('discount', 0))
        product['sku'] = str(product.get('sku', 'MISSING'))
    
    for customer in data['customers']:
        customer['customer_name'] = str(customer.get('customer_name', 'MISSING'))
        customer['phone_number'] = str(customer.get('phone_number', 'MISSING'))
        customer['total_purchase_amount'] = safe_float(customer.get('total_purchase_amount', 0))
        customer['email'] = str(customer.get('email', 'MISSING'))
        customer['address'] = str(customer.get('address', 'MISSING'))
    
    summary = data['summary']
    summary['total_quantity'] = safe_int(summary.get('total_quantity', 0))
    summary['total_amount'] = safe_float(summary.get('total_amount', 0))
    summary['cgst'] = safe_float(summary.get('cgst', 0))
    summary['sgst'] = safe_float(summary.get('sgst', 0))
    summary['igst'] = safe_float(summary.get('igst', 0))
    summary['net_amount'] = safe_float(summary.get('net_amount', 0))
    summary['total_tax'] = safe_float(summary.get('total_tax', 0))
    summary['extra_discount'] = safe_float(summary.get('extra_discount', 0))
    summary['round_off'] = safe_float(summary.get('round_off', 0))
    
    return data


def format_date(date_value) -> str:
    """Format date"""
    if not date_value or date_value == 'None':
        return 'MISSING'
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    return str(date_value)


def safe_float(value) -> float:
    """Safely convert to float"""
    if value is None or value == '' or value == 'None':
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(',', '')
        return float(value)
    except:
        return 0.0


def safe_int(value) -> int:
    """Safely convert to int"""
    if value is None or value == '' or value == 'None':
        return 0
    try:
        if isinstance(value, str):
            value = value.replace(',', '')
        return int(float(value))
    except:
        return 0